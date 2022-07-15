import datetime
import os

import openpyxl
from django.http import HttpResponse
from openpyxl import load_workbook

from django.core.paginator import Paginator
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Q

from .models import Agent, Bso, PolicyAgents, HistoryBso, StatusBso, Channel, NameFinancial, Financial, InputCom
from main.models import Company, Type

from main.views import get_start_end_date


@login_required(login_url='login')
def issuance_bso(request):
    agents = Agent.objects.all()
    company = Company.objects.all()
    status_bso = StatusBso.objects.all()
    bso_for_agents = Bso.objects.filter(clear=True)
    errors = {}
    errors_unload = ''
    errors_count = ''
    number_for_errors = 0
    selected = {}
    date_now = datetime.datetime.now().strftime("%Y-%m-%d")
    text_bso = ''
    text_errors_bso = 0
    if request.method == 'POST':
        # Создание статуса бланка
        if 'add_status_bso' in request.POST:
            StatusBso.objects.get_or_create(name=request.POST['status_name'])

        # Передача БСО со склада агенту
        if 'bso_in_agents' in request.POST:
            for bso_id in request.POST:
                if 'check_bso' in bso_id:
                    bso_for_edit = Bso.objects.get(id=request.POST[bso_id])
                    bso_for_edit.agent = Agent.objects.get(id=request.POST['agent'])
                    bso_for_edit.clear = False
                    bso_for_edit.save()

                    status, created = StatusBso.objects.get_or_create(name='Выдан агенту')
                    HistoryBso(
                        date_at=datetime.datetime.now().date(),
                        status=status,
                        bso=bso_for_edit,
                    ).save()

        # Добавление новых БСО на склад вручную
        if 'add_new_bso' in request.POST:
            for i in range(int(request.POST['range'])):
                new_bso, created = Bso.objects.get_or_create(
                    series=request.POST['series'],
                    number=int(request.POST['number']) + i,
                    company_id=request.POST['company'],
                    defaults={
                        'date_add': request.POST['date_add'],
                    }
                )

                if not created:
                    errors[f'{new_bso.series} {new_bso.number}'] = 'Уже есть в системе'

                else:
                    status, created = StatusBso.objects.get_or_create(name='Чистый')
                    HistoryBso(
                        date_at=datetime.datetime.now().date(),
                        status=status,
                        bso=new_bso,
                    ).save()

                if len(errors) > 0:
                    # Выгрузка незагруженных БСО в xlsx
                    pass

        # Загрузка БСО из файла
        if 'add_new_bso_from_file' in request.POST:
            errors = {}
            wb = load_workbook(filename=request.FILES['file'])
            sheet = wb.worksheets[0]
            i = 2
            while True:
                if sheet[i][0].value is None or sheet[i][0].value == '':
                    break
                company_for_new_bso, created = Company.objects.get_or_create(name=sheet[i][2].value)
                new_bso, created = Bso.objects.get_or_create(
                    series=sheet[i][0].value,
                    number=sheet[i][1].value,
                    company=company_for_new_bso,
                    defaults={
                        'date_add': datetime.datetime.now().date(),
                    }
                )

                if not created:
                    errors[f'{new_bso.series} {new_bso.number}'] = 'Уже есть в системе'
                else:
                    status, created = StatusBso.objects.get_or_create(name='Чистый')
                    HistoryBso(
                        date_at=datetime.datetime.now().date(),
                        status=status,
                        bso=new_bso,
                    ).save()

                i += 1

            if len(errors) > 0:
                # Выгрузка незагруженных БСО в xlsx
                pass

        # Смена статуса БСО из файла
        if 'change_status_from_file' in request.POST:
            errors = {}
            wb = load_workbook(filename=request.FILES['file'])
            sheet = wb.worksheets[0]
            i = 2
            while True:
                series = ''
                number = ''
                # если серия и номер пустые, прекращяем читать файл
                if sheet[i][0].value is None or sheet[i][0].value == '':
                    if sheet[i][1].value is None or sheet[i][1].value == '':
                        break

                if sheet[i][0].value is not None and sheet[i][0].value != '':
                    series = sheet[i][0].value
                if sheet[i][1].value is not None and sheet[i][1].value != '':
                    number = sheet[i][1].value

                bso_for_change = Bso.objects.filter(
                    series=series,
                    number=number,
                )

                if len(bso_for_change) > 0:
                    number_for_errors += 1
                    new_status = StatusBso.objects.get(id=request.POST['status_for_change'])

                    if new_status.name == 'Чистый':
                        bso_for_change[0].clear = True
                        bso_for_change[0].save()
                    else:
                        bso_for_change[0].clear = False
                        bso_for_change[0].save()

                    HistoryBso(
                        date_at=datetime.datetime.now().date(),
                        status=new_status,
                        bso=bso_for_change[0],
                    ).save()
                    policy_for_edit = PolicyAgents.objects.filter(bso=bso_for_change[0])
                    if len(policy_for_edit) > 0:
                        policy_for_edit[0].status = new_status
                        policy_for_edit[0].save()
                else:
                    errors[f'{series} {number}'] = 'БСО не найден'

                i += 1

            # выгрузка не найденных БСО
            if len(errors) > 0:
                errors_unload = 1
                errors_count = len(errors)

                wb = openpyxl.Workbook()
                sheet = wb['Sheet']

                sheet['A1'] = '№'
                sheet['B1'] = 'БСО'
                sheet['C1'] = 'Тип ошибки'

                wb.save('agents/file/errors.xlsx')

                wb = openpyxl.load_workbook('agents/file/errors.xlsx')
                sheet = wb['Sheet']

                str_number = 2
                for error, text in errors.items():
                    sheet[str_number][0].value = str_number - 1
                    sheet[str_number][1].value = error
                    sheet[str_number][2].value = text
                    str_number += 1

                wb.save('agents/file/errors.xlsx')

        # Точечная смена статуса БСО
        if 'bso_id' in request.POST:
            bso_for_edit = Bso.objects.get(id=int(request.POST['bso_id']))
            new_status = StatusBso.objects.get(id=request.POST['status_for_change'])

            if new_status.name == 'Чистый':
                bso_for_edit.clear = True
                bso_for_edit.save()
            else:
                bso_for_edit.clear = False
                bso_for_edit.save()

            HistoryBso(
                date_at=datetime.datetime.now().date(),
                status=new_status,
                bso=bso_for_edit,
            ).save()

            policy_for_edit = PolicyAgents.objects.filter(bso=bso_for_edit)
            if len(policy_for_edit) > 0:
                policy_for_edit[0].status = new_status
                policy_for_edit[0].save()

    bso = Bso.objects.all()

    text_search = ''
    if 'search' in request.GET:
        bso = bso.filter(number__icontains=request.GET.get('search'))
        text_search = f'Найдено {len(bso)} БСО'

    else:
        if 'agent' in request.GET and request.GET.get('agent') != 'all':
            # Сортировка по агенту
            bso = bso.filter(agent_id=request.GET.get('agent'))
            selected['agent'] = int(request.GET.get('agent'))

        if 'status_bso' in request.GET:
            # Сортировка по статусу БСО
            if request.GET['status_bso'] != 'all':
                selected['status_bso'] = int(request.GET['status_bso'])
                status_for_filter = StatusBso.objects.get(id=request.GET['status_bso'])
                bso_list = []
                for b in bso:
                    # Выборка актуального статуса
                    old_history = HistoryBso.objects.filter(bso_id=b.id)[0]
                    for his in HistoryBso.objects.filter(bso_id=b.id):
                        if his.id > old_history.id:
                            old_history = his
                    if status_for_filter == old_history.status:
                        bso_list.append(b)
                bso = bso_list

    bso_list = []
    for b in bso:
        temp_dict = {}
        temp_dict['id'] = b.id
        temp_dict['company'] = b.company.name
        temp_dict['series'] = b.series
        temp_dict['number'] = b.number
        temp_dict['date_add'] = b.date_add

        old_history = HistoryBso.objects.filter(bso_id=b.id)[0]

        # Выборка актуального статуса
        for his in HistoryBso.objects.filter(bso_id=b.id):
            if his.id > old_history.id:
                old_history = his

        temp_dict['history_status'] = old_history.status
        temp_dict['history_date_at'] = old_history.date_at
        temp_dict['history'] = HistoryBso.objects.filter(bso_id=b.id)
        temp_dict['agent'] = b.agent
        bso_list.append(temp_dict)

    # ссылка с параметрами для пагинации
    link = '?'
    for key, value in request.GET.items():
        if key == 'page' or key == 'page_2':
            continue
        link = link + f'{key}={value}&'

    # Пагинация основного списка БСО
    paginator = Paginator(bso_list, 15)
    current_page = request.GET.get('page', 1)
    page = paginator.get_page(current_page)

    # Пагинация списка БСО для передачи агенту
    paginator_2 = Paginator(bso_for_agents, 20)
    current_page_2 = request.GET.get('page_2', 1)
    page_2 = paginator_2.get_page(current_page_2)

    context = {
        'agents': agents,
        'number_for_errors': number_for_errors,
        'errors_count': errors_count,
        'errors_unload': errors_unload,
        'status_bso': status_bso,
        'date_now': date_now,
        'selected': selected,
        'company': company,
        'text_bso': text_bso,
        'text_errors_bso': text_errors_bso,
        'errors': errors,
        'bso_list': page.object_list,
        'bso_for_agents': page_2.object_list,
        'text_search': text_search,
        'page': page,
        'page_2': page_2,
        'link': link,
        'paginator': paginator,
        'paginator_2': paginator_2,
    }
    return render(request, 'agents/issuance_bso.html', context)


@login_required(login_url='login')
def receivable(request):
    text = ''
    text_errors = 0
    errors_count = ''
    errors = {}
    number = ''
    errors_unload = 0
    selected = {}
    text_agent = ''
    response = ''
    type = Type.objects.all()
    agents = Agent.objects.all()
    status = StatusBso.objects.all()
    channels = Channel.objects.all()

    if request.method == 'POST':
        # добавление агента
        if 'name' in request.POST:
            agent, created = Agent.objects.get_or_create(
                name=request.POST.get('name'),
                defaults={
                    'storage_time': request.POST.get('storage_time')
                }
            )

            if created:
                text_agent = 'Агент добавлен'
            else:
                text_agent = 'Агент с таким именем уже есть в системе'

        # первый этап загрузки агентских продаж из файла
        if 'upload' in request.POST:
            headers = {}

            headers_for_upload = {
                'Полис': 'policy',
                'Тип полиса': 'type',
                'Страхователь': 'client',
                'Тип страхователя': 'type_client',
                'Дата подписания договора': 'date_at',
                'Начало': 'date_start',
                'Окончание': 'date_end',
                'Общая премия': 'sp',
                'Тип оплаты': 'type_pay',
                'Страховая компания': 'company',
                'Канал продаж': 'channel',
            }

            wb = openpyxl.load_workbook(filename=request.FILES['file'])
            sheet = wb.worksheets[0]

            col = 0
            step_upload = 1
            while True:
                try:
                    if sheet[1][col].value is None or sheet[1][col].value == '':
                        break
                    headers[col] = sheet[1][col].value
                    col += 1

                except:
                    break

            wb.save('agents/file/upload.xlsx')

            context = {
                'headers': headers,
                'agents': agents,
                'step_upload': step_upload,
                'headers_for_upload': headers_for_upload,
            }

            return render(request, 'agents/receivable.html', context)

        # второй этап загрузки агентских продаж из файла
        if 'step_upload' in request.POST:
            wb = openpyxl.load_workbook('agents/file/upload.xlsx')
            sheet = wb.worksheets[0]

            number = 0
            row = 2
            while True:
                if sheet[row][0].value is None or sheet[row][0].value == '':
                    break

                if not isinstance(sheet[row][int(request.POST['date_at'])].value, datetime.date):
                    errors[sheet[row][int(request.POST['policy'])].value] = 'Дата подписания - нужен формат даты'
                    row += 1
                    continue
                if not isinstance(sheet[row][int(request.POST['date_start'])].value, datetime.date):
                    errors[sheet[row][int(request.POST['policy'])].value] = 'Начало - нужен формат даты'
                    row += 1
                    continue
                if not isinstance(sheet[row][int(request.POST['date_end'])].value, datetime.date):
                    errors[sheet[row][int(request.POST['policy'])].value] = 'Окончание - нужен формат даты'
                    row += 1
                    continue

                channel = Channel.objects.get_or_create(name=sheet[row][int(request.POST['channel'])].value)[0]
                company = Company.objects.get_or_create(name=sheet[row][int(request.POST['company'])].value)[0]
                new_type = Type.objects.get_or_create(name=sheet[row][int(request.POST['type'])].value)[0]
                agent = Agent.objects.get(id=request.POST['agent'])
                new_policy, created = PolicyAgents.objects.get_or_create(
                    policy=sheet[row][int(request.POST['policy'])].value,
                    company=company,
                    defaults={
                        'type': new_type,
                        'channel': channel,
                        'agent': agent,
                        'client': sheet[row][int(request.POST['client'])].value,
                        'type_client': sheet[row][int(request.POST['type_client'])].value,
                        'date_registration': sheet[row][int(request.POST['date_at'])].value,
                        'date_start': sheet[row][int(request.POST['date_start'])].value,
                        'date_end': sheet[row][int(request.POST['date_end'])].value,
                        'price': float(sheet[row][int(request.POST['sp'])].value),
                        'type_pay': sheet[row][int(request.POST['type_pay'])].value,
                    }
                )

                if created:
                    number += 1
                    # Если полис электронный
                    if 'б' in new_policy.type_pay.lower():
                        new_status = StatusBso.objects.get_or_create(name='Согласование в СК')[0]
                        new_policy.status = new_status
                        new_policy.save()
                    else:
                        # связка с бсо
                        policy_bso = new_policy.policy
                        # убираем все пробелы с конца
                        while policy_bso[-1] == ' ':
                            policy_bso = policy_bso[:-1]
                        # если есть пробел, разбиваем серию и номер пробелом
                        if ' ' in policy_bso:
                            new_series = policy_bso.split()[0]
                            new_number = policy_bso.split()[1]
                        else:
                            if 'осаго' in new_policy.type.name.lower():
                                new_series = policy_bso[:3]
                                new_number = policy_bso[3:]
                            else:
                                new_series = ''
                                new_number = policy_bso
                        new_bso = Bso.objects.filter(series=new_series, number=new_number)
                        print(new_bso)
                        if len(new_bso) > 0:
                            new_policy.bso = new_bso[0]
                            new_policy.save()

                            new_status = StatusBso.objects.get_or_create(name='АКТ приёма-передачи')[0]
                            new_policy.status = new_status
                            new_policy.save()
                            HistoryBso(
                                date_at=datetime.datetime.now().date(),
                                status=new_status,
                                bso=new_bso[0],
                            ).save()
                else:
                    errors[sheet[row][int(request.POST['policy'])].value] = 'Полис уже есть в базе'
                    pass

                row += 1

            os.remove('agents/file/upload.xlsx')

            # сохраняем файл с ошибками загрузки
            if len(errors) > 0:
                errors_unload = 1
                errors_count = len(errors)

                wb = openpyxl.Workbook()
                sheet = wb['Sheet']

                sheet['A1'] = '№'
                sheet['B1'] = 'БСО'
                sheet['C1'] = 'Тип ошибки'

                str_number = 2
                for error, text in errors.items():
                    sheet[str_number][0].value = str_number - 1
                    sheet[str_number][1].value = error
                    sheet[str_number][2].value = text
                    str_number += 1

                wb.save('agents/file/errors.xlsx')

    # дата начала и окончания периода выборки
    if 'date_start' not in request.GET:
        date_start, date_end = get_start_end_date()
        date_start = date_start.strftime("%Y-%m-%d")
        date_end = date_end.strftime("%Y-%m-%d")
    else:
        date_start = request.GET.get('date_start')
        date_end = request.GET.get('date_end')

    if 'search' in request.GET:
        # поиск по страхователю или по номеру полиса
        policy = PolicyAgents.objects.filter(
            Q(client__iregex=request.GET.get('search')) |
            Q(policy__iregex=request.GET.get('search'))
        )

    else:
        # отбор полисов, оформленных за период (текущий месяц по умолчанию)
        policy = PolicyAgents.objects.filter(
            date_registration__lt=date_end,
            date_registration__gte=date_start,
        )

        # Фильтр по статусу полиса
        if 'status' in request.GET and request.GET.get('status') != 'all':
            selected['status'] = int(request.GET.get("status"))
            policy = policy.filter(status_id=request.GET.get("status"))

        # Фильтр по агенту
        if 'agent' in request.GET and request.GET.get('agent') != 'all':
            selected['agent'] = int(request.GET.get("agent"))
            policy = policy.filter(agent_id=request.GET.get("agent"))

        # Фильтр по каналу продаж
        if 'channel' in request.GET and request.GET['channel'] != 'all':
            selected['channel'] = int(request.GET['channel'])
            policy = policy.filter(channel_id=request.GET['channel'])

    # ссылка с параметрами для пагинации
    link = '?'
    for key, value in request.GET.items():
        if key == 'page':
            continue
        link = link + f'{key}={value}&'

    # пагинация результата
    paginator = Paginator(policy, 15)
    current_page = request.GET.get('page', 1)
    page = paginator.get_page(current_page)

    context = {
        'agents': agents,
        'number': number,
        'status': status,
        'errors_unload': errors_unload,
        'errors_count': errors_count,
        'response': response,
        'channels': channels,
        'type': type,
        'policy': page.object_list,
        'date_start': date_start,
        'date_end': date_end,
        'text': text,
        'text_errors': text_errors,
        'errors': errors,
        'selected': selected,
        'text_agent': text_agent,
        'page': page,
        'link': link,
        'paginator': paginator,
    }

    return render(request, 'agents/receivable.html', context)


@login_required(login_url='login')
def unload_receivable(request):
    # выгрузка агентских продаж в xlsx
    if request.user.admin:
        if request.method == 'POST':
            policy = PolicyAgents.objects.filter(
                date_registration__gte=request.POST['date_start'],
                date_registration__lt=request.POST['date_end'],
            )

            if request.POST['status'] != 'all':
                policy = policy.filter(status=request.POST['status'])
            if request.POST['agent'] != 'all':
                policy = policy.filter(agent_id=request.POST['agent'])
            if request.POST['type'] != 'all':
                policy = policy.filter(type_id=request.POST['type'])

            wb = openpyxl.Workbook()
            sheet = wb['Sheet']

            sheet['A1'] = '№'
            sheet['B1'] = 'Агент'
            sheet['C1'] = 'Тип полиса'
            sheet['D1'] = 'Полис'
            sheet['E1'] = 'Клиент'
            sheet['F1'] = 'Тип клиента'
            sheet['G1'] = 'Комиссия агента в %'
            sheet['H1'] = 'Комиссия агента в руб.'
            sheet['I1'] = 'Дата подписания договора'
            sheet['J1'] = 'Дата начала действия'
            sheet['K1'] = 'Дата окончания действия'
            sheet['L1'] = 'Общая премия'
            sheet['M1'] = 'Тип оплаты'
            sheet['N1'] = 'Статус полиса'

            wb.save('agents/file/receivable.xlsx')

            wb = openpyxl.load_workbook('agents/file/receivable.xlsx')
            sheet = wb['Sheet']

            str_number = 2
            for policy in policy:
                sheet[str_number][0].value = str_number - 1
                sheet[str_number][1].value = policy.agent.name
                sheet[str_number][2].value = policy.type.name
                sheet[str_number][3].value = policy.policy
                sheet[str_number][4].value = policy.client
                sheet[str_number][5].value = policy.type_client
                sheet[str_number][6].value = policy.agent_commission
                sheet[str_number][7].value = policy.agent_commission_rub
                sheet[str_number][8].value = policy.date_registration
                sheet[str_number][9].value = policy.date_start
                sheet[str_number][10].value = policy.date_end
                sheet[str_number][11].value = policy.price
                sheet[str_number][12].value = policy.type_pay
                if policy.status == 'receivable':
                    sheet[str_number][13].value = 'Дебиторка'
                elif policy.status == 'reconciliation':
                    sheet[str_number][13].value = 'Сверки'
                elif policy.status == 'accept':
                    sheet[str_number][13].value = 'Акцепт'

                str_number += 1

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="receivable.xlsx"'

            wb.save(response)

            return response


@login_required(login_url='login')
def unload_errors(request):
    # выгрузка незагруженных полисов
    wb = openpyxl.load_workbook('agents/file/errors.xlsx')

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="receivable.xlsx"'

    wb.save(response)

    return response


@login_required(login_url='login')
def financial(request):
    # финансовая политика
    date_now = datetime.datetime.now().date().strftime("%Y-%m-%d")
    company_dict = ''
    fin_name = ''
    name_id = ''
    company_dict_for_edit = ''
    com_count = 0

    if request.method == 'POST':
        # добавление входящего кв
        if 'add_input_com' in request.POST:
            new_input_com, created = InputCom.objects.get_or_create(
                company_id=request.POST['company'],
                channel_id=request.POST['channel'],
                date_start=request.POST['date_start'],
                type_policy_id=request.POST['type_policy'],
                defaults={
                    'value': request.POST['input_com_value'],
                }
            )

            if not created:
                new_input_com.value = request.POST['input_com_value']
                new_input_com.save()

        # добавление фин. сетки
        if 'add_name_financial' in request.POST:
            new_name_financial, created = NameFinancial.objects.get_or_create(
                name=request.POST['name_financial'],
                date_start=request.POST['date_start'],
            )

        # редактирование фин. сетки
        if 'name_id' in request.POST:
            com_count = 2
            for value in request.POST:
                if 'value' not in value:
                    continue
                values = value.split("_")

                if float(request.POST[value].replace(',', '.')) == 0:
                    fin_for_del = Financial.objects.filter(
                        name_id=request.POST['name_id'],
                        company_id=values[1],
                        channel_id=values[2],
                        type_policy_id=values[3],
                    )
                    if len(fin_for_del) > 0:
                        fin_for_del.delete()
                    continue

                new_financial, created = Financial.objects.get_or_create(
                    name_id=request.POST['name_id'],
                    company_id=values[1],
                    channel_id=values[2],
                    type_policy_id=values[3],
                    defaults={
                        'agent_com': float(request.POST[value].replace(',', '.')),
                    }
                )

                if not created:
                    new_financial.agent_com = float(request.POST[value].replace(',', '.'))
                    new_financial.save()

    input_com = InputCom.objects.all()
    name_financial = NameFinancial.objects.all()
    company = Company.objects.all()
    type_policy = Type.objects.all()
    channel = Channel.objects.all()

    # выгрузка КВ агентов
    if 'agent_count' in request.GET:
        com_count = 1

    # подготовка к редактированию фин. сетки
    if 'name_for_edit' in request.POST:
        com_count = 2
        name_id = request.POST['name_for_edit']

    # сортировка для вывода входящего КВ
    if com_count == 0:
        # все компании, по которым есть вход кв
        inp_company_ids = set(com.company.id for com in input_com)

        # сортировка: "компания"-"канал продаж"-"дата начала"-"тип полиса: кв"
        company_dict = {}
        for company_id in inp_company_ids:
            # список всех каналов продаж для конкретной ск по которым есть вход кв
            input_com_company = input_com.filter(company_id=company_id)
            inp_channel_ids = set(com.channel.id for com in input_com_company)

            company_name = Company.objects.get(id=company_id)
            channel_dict = {}
            for channel_id in inp_channel_ids:
                # разбивка по датам начала действия вход кв конкретного канала продаж конкретной ск
                input_com_channel = input_com_company.filter(channel_id=channel_id)
                inp_date = set(com.date_start for com in input_com_channel)
                inp_date = sorted(list(inp_date), reverse=True)

                channel_name = Channel.objects.get(id=channel_id)
                date_dict = {}
                for date_start in inp_date:
                    # отбор по дате начала действия
                    inp_com_values = input_com_channel.filter(date_start=date_start)

                    temp_dict = {}
                    for com_values in inp_com_values:
                        # итоговые значения кв
                        temp_dict[com_values.type_policy.name] = com_values.value
                    date_dict[date_start] = temp_dict
                channel_dict[channel_name.name] = date_dict
            company_dict[company_name.name] = channel_dict

    # сортировка для вывода фин. сеток
    if com_count == 1:
        # все финансовые сетки
        name_financial_ids = set(name_fin.id for name_fin in name_financial)

        # сортировка: "фин. сетка"-"компания"-"канал продаж"-"дата начала"-"тип полиса: кв"
        fin_name = {}
        for name_fin_id in name_financial_ids:
            # список всех компаний, которые добавленны в конкретную сетку
            financial_for_name = Financial.objects.filter(name_id=name_fin_id)
            fin_company_ids = set(fin.company.id for fin in financial_for_name)

            financial_name = NameFinancial.objects.get(id=name_fin_id)
            fin_company = {}
            for company_id in fin_company_ids:
                # список всех каналов продаж по каждой компании из списка
                financial_for_company = financial_for_name.filter(company_id=company_id)
                fin_channel_id = set(fin.channel.id for fin in financial_for_company)

                name_company = Company.objects.get(id=company_id)
                fin_channel = {}
                for channel_id in fin_channel_id:
                    # итоговые значения
                    financial_for_channel = financial_for_company.filter(channel_id=channel_id)

                    name_channel = Channel.objects.get(id=channel_id)
                    fin_channel[name_channel] = financial_for_channel
                fin_company[name_company] = fin_channel
            fin_name[financial_name] = fin_company

    # сортировка сеток для редактирования
    if com_count == 2:
        # все компании, по которым есть вход кв
        inp_company_ids = set(com.company.id for com in input_com)

        # сортировка: "компания"-"канал продаж"-"тип полиса"
        company_dict_for_edit = {}
        for company_id in inp_company_ids:
            # список всех каналов продаж для конкретной ск по которым есть вход кв
            input_com_company = input_com.filter(company_id=company_id)
            inp_channel_ids = set(com.channel.id for com in input_com_company)

            company_name = Company.objects.get(id=company_id)
            channel_dict_for_edit = {}
            for channel_id in inp_channel_ids:
                # разбивка по типу полисов конкретного канала продаж конкретной ск
                input_com_channel = input_com_company.filter(channel_id=channel_id)
                type_policy_for_edit = {
                    input_c.type_policy: input_c.type_policy.id for input_c in input_com_channel
                }

                # проставляем текущие значения кв агента в текущей сетке, если такого нет, то = 0
                for type_name, type_id in type_policy_for_edit.items():
                    financial_for_edit = Financial.objects.filter(
                        name_id=name_id,
                        company_id=company_id,
                        channel_id=channel_id,
                        type_policy_id=type_id,
                    )

                    if len(financial_for_edit) > 0:
                        type_policy_for_edit[type_name] = financial_for_edit[0].agent_com
                    else:
                        type_policy_for_edit[type_name] = 0

                channel_name = Channel.objects.get(id=channel_id)
                channel_dict_for_edit[channel_name] = type_policy_for_edit
            company_dict_for_edit[company_name] = channel_dict_for_edit

    context = {
        'company_dict': company_dict,
        'company_dict_for_edit': company_dict_for_edit,
        'fin_name': fin_name,
        'name_id': name_id,
        'com_count': com_count,
        'date_now': date_now,
        'company': company,
        'type_policy': type_policy,
        'channel': channel,
    }

    return render(request, 'agents/financial.html', context)
