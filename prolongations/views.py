import datetime
import locale
import csv
import re
from PyPDF2 import PdfReader

from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.db.models import Q
from openpyxl import load_workbook

from .forms import UploadFileForm

from .models import Status, PolicyBase, NameBase
from main.models import Client, Type, Company, Commission, Policy, Channel


# Установка родной локации для вывода кириллицей
locale.setlocale(locale.LC_ALL, "")


# Месяцы
months_rus = {
    '01': 'Январь',
    '02': 'Февраль',
    '03': 'Март',
    '04': 'Апрель',
    '05': 'Май',
    '06': 'Июнь',
    '07': 'Июль',
    '08': 'Август',
    '09': 'Сентябрь',
    '10': 'Октябрь',
    '11': 'Ноябрь',
    '12': 'Декабрь',
}


def pars_pdf_ingos(file):
    # парсинг инфы с полиса ингоса осаго pdf
    months = {
        'января': '01',
        'февраля': '02',
        'марта': '03',
        'апреля': '04',
        'мая': '05',
        'июня': '06',
        'июля': '07',
        'августа': '08',
        'сентября': '09',
        'октября': '10',
        'ноября': '11',
        'декабря': '12',
    }

    reader = PdfReader(file)
    text = reader.pages[0].extractText()
    series_number = re.search('№ \\w{3} {1,}\\d{10}', text)
    if series_number is not None:
        series = re.search('№ \\w{3} {1,}\\d{10}', text).group(0)[2:5]
        number = re.search('№ \\w{3} {1,}\\d{10}', text).group(0)[-10:]
    else:
        return 'Ошибка: Серия и Номер не распознаны'

    sp = re.search('\\d{0,6} руб\. \\d{0,2}', text)
    if sp is not None:
        sp = re.search('\\d{0,6} руб\. \\d{0,2}', text).group(0).replace(" руб. ", ".")
    else:
        return 'Ошибка: Премия не распознана'

    dates = re.finditer('\\d{2}.\\d{2}.\\d{4} г.', text)
    for i, date in enumerate(dates):
        if i == 0:
            date_start = date.group(0)[:-3]
            date_start = f'{date_start[-4:]}-{date_start[3:5]}-{date_start[:2]}'
        elif i == 1:
            date_end = date.group(0)[:-3]
            date_end = f'{date_end[-4:]}-{date_end[3:5]}-{date_end[:2]}'
    date_registration = re.search('Дата заключения договора .\\d{1,2}. \\w{3,8} {1,}\\d{4}', text).group(0)
    if date_registration is not None:
        date_registration = re.search(
            'Дата заключения договора .\\d{1,2}. \\w{3,8} {1,}\\d{4}', text).group(0).replace(" ", "")
        date_registration = f'{date_registration[-4:]}-{months[date_registration[26:-4]]}-{date_registration[23:25]}'
    else:
        return 'Ошибка: Дата оформления не распознанна'

    client = re.search('Страхователь:\\s.{1,}\\s', text)
    if client is not None:
        client = re.search('Страхователь:\\s.{1,}\\s', text).group(0)[14:-1].replace("  ", " ")
    else:
        return 'Ошибка: Страховател не распознан'

    vin = re.search('\\w( \\w){16}', text)
    if vin is not None:
        vin = vin.group(0).replace(" ", "")
    else:
        vin = re.search('\\w-\\w\\s\\w( \\w){14}', text)
        if vin is not None:
            vin = vin.group(0).replace(" ", "").replace('\n', "").replace('-', "")
        else:
            return 'Ошибка: VIN не распознан'

    result = {}
    result['client_full_name'] = client
    result['series'] = series
    result['number'] = number
    result['sp'] = sp
    result['date_start'] = date_start
    result['date_end'] = date_end
    result['date_registration'] = date_registration
    result['vin'] = vin

    return result


@login_required(login_url='login')
def upload_policy(request):
    # Изменение статуса полиса
    if request.user.agent == False and request.user.admin == True or request.user.username == 'APuchkova':
        if request.method == 'POST':
            policy_up = PolicyBase.objects.get(id=request.POST.get('policy'))
            policy_up.status = Status.objects.get(id=request.POST.get('status'))
            policy_up.save()

        return HttpResponseRedirect('/prolongations/status_change')


@login_required(login_url='login')
def status_change(request):
    # Статистика по пролонгации
    error_count = 0
    error = ''
    if request.user.agent == False and request.user.admin == True or request.user.username == 'APuchkova':
        text = ''
        form = UploadFileForm()
        status = Status.objects.all()
        policy_base = PolicyBase.objects.filter(status__name='В работе').order_by('date_end')

        if 'search' in request.GET:
            policy_base = policy_base.filter(
                Q(client__first_name__iregex=request.GET.get('search')) |
                Q(client__middle_name__iregex=request.GET.get('search')) |
                Q(client__last_name__iregex=request.GET.get('search'))
            )

        if request.method == 'POST':
            # загрузка из файла
            if 'upload' in request.POST:
                form = UploadFileForm(request.POST, request.FILES)
                if form.is_valid():
                    wb = load_workbook(filename=request.FILES['file'])
                    sheet = wb.worksheets[0]

                    for row in range(2, sheet.max_row + 1):
                        # пропускаем пустую строчку
                        if sheet[row][0].value is None:
                            continue

                        type_policy, created = Type.objects.get_or_create(name=sheet[row][10].value)
                        name_base, created = NameBase.objects.get_or_create(name=sheet[row][11].value)
                        company_policy, created = Company.objects.get_or_create(name=sheet[row][4].value)

                        new_full_name = sheet[row][0].value

                        while new_full_name[-1] == ' ':
                            new_full_name = new_full_name[:-1]
                        last_name = new_full_name.split()[0].capitalize()
                        first_name = new_full_name.split()[1].capitalize()
                        middle_name = ''
                        if len(new_full_name.split()) > 2:
                            for name in new_full_name.split()[2:]:
                                middle_name = middle_name + name.capitalize() + " "
                            middle_name = middle_name[:-1]

                        client_policy, created = Client.objects.get_or_create(
                            first_name=first_name,
                            last_name=last_name,
                            middle_name=middle_name,
                            phone=sheet[row][2].value,
                            defaults={
                                'email': sheet[row][3].value,
                                'birthday': sheet[row][1].value,
                            }
                        )

                        status_policy = Status.objects.get(name='В работе')
                        bso_policy = sheet[row][5].value
                        sp_policy = float(str(sheet[row][9].value).replace(',', '.'))
                        channel_policy = sheet[row][12].value
                        object_policy = f'{sheet[row][6].value} {sheet[row][7].value} {sheet[row][8].value}'
                        manager_policy = sheet[row][13].value
                        date_end = sheet[row][14].value

                        PolicyBase.objects.get_or_create(
                            bso=bso_policy,
                            company=company_policy,
                            defaults={
                                'type': type_policy,
                                'name': name_base,
                                'client': client_policy,
                                'status': status_policy,
                                'sp': sp_policy,
                                'channel': channel_policy,
                                'manager': manager_policy,
                                'object': object_policy,
                                'date_end': date_end,
                            }
                        )

            # ручное добавление полиса
            if 'add_new_policy' in request.POST:
                client = Client.objects.get(id=request.POST['client_id'])
                if client.phone != request.POST['client_phone']:
                    client.phone = request.POST['client_phone']
                    client.save()
                if client.email != request.POST['client_email']:
                    client.email = request.POST['client_email']
                    client.save()

                company = Company.objects.get(name=request.POST['policy_company'])
                channel = Channel.objects.get(name=request.POST['policy_channel'])
                type = Type.objects.get(name=request.POST['policy_type'])

                policy, created = Policy.objects.get_or_create(
                    number=request.POST['policy_number'],
                    series=request.POST['policy_series'],
                    company_id=company.id,
                    defaults={
                        'user_id': request.user.id,
                        'client_id': client.id,
                        'channel_id': channel.id,
                        'type_id': type.id,
                        'date_registration': request.POST['date_registration'],
                        'date_start': request.POST['date_start'],
                        'date_end': request.POST['date_end'],
                        'commission': 0,
                        'sp': float(request.POST['sp'].replace(',', '.')),
                        'status': 'prolongation',
                    }
                )

                if created:
                    commission_objects = Commission.objects.filter(
                        type=policy.type,
                        channel=policy.channel,
                        company=policy.company,
                        date_start__lte=policy.date_registration,
                    )

                    if len(commission_objects) > 0:
                        policy.commission = commission_objects.order_by('-date_start')[0].value
                        policy.save()

                policy_up = PolicyBase.objects.get(id=request.POST['policy_up'])
                policy_up.status = Status.objects.get(id=request.POST['status'])
                policy_up.save()

            # проведение из загруженного файла
            if 'upload_file' in request.POST:
                policy_ = pars_pdf_ingos(request.FILES['file'])
                if 'Ошибка' in policy_:
                    error = policy_
                    error_count = 1
                else:
                    policy_up = PolicyBase.objects.get(id=request.POST['policy_up'])
                    if policy_['client_full_name'].lower() not in str(policy_up.client).lower():
                        error = 'Страхователь в полисе не совпадает с выбранным клиентом'
                        error_count = 1
                    elif policy_['vin'].lower() not in policy_up.object.lower():
                        error = 'VIN в полисе не совпадает с текущим полисом'
                        error_count = 1
                    else:
                        policy_up.status = Status.objects.get(id=request.POST['status'])
                        policy_up.save()

                        company = Company.objects.get(name='Ингосстрах')
                        channel = Channel.objects.get(name='ООО "НьюЛукИншуранс"')
                        type = Type.objects.get(name='ОСАГО')
                        client = Client.objects.get(id=request.POST['client_id'])

                        policy, created = Policy.objects.get_or_create(
                            number=policy_['number'],
                            series=policy_['series'],
                            company_id=company.id,
                            defaults={
                                'user_id': request.user.id,
                                'client_id': client.id,
                                'channel_id': channel.id,
                                'type_id': type.id,
                                'date_registration': policy_['date_registration'],
                                'date_start': policy_['date_start'],
                                'date_end': policy_['date_end'],
                                'commission': 0,
                                'sp': float(policy_['sp']),
                                'status': 'prolongation',
                            }
                        )

                        if created:
                            commission_objects = Commission.objects.filter(
                                type=policy.type,
                                channel=policy.channel,
                                company=policy.company,
                                date_start__lte=policy.date_registration,
                            )

                            if len(commission_objects) > 0:
                                policy.commission = commission_objects.order_by('-date_start')[0].value
                                policy.save()

        paginator = Paginator(policy_base, 20)
        current_page = request.GET.get('page', 1)
        page = paginator.get_page(current_page)

        context = {
            'text': text,
            'error': error,
            'error_count': error_count,
            'form': form,
            'page': page,
            'policy_base': page.object_list,
            'paginator': paginator,
            'status': status,
        }

        return render(request, 'prolongations/status_change.html', context)


@login_required(login_url='login')
def new_statistics(request):
    # Статистика по пролонгационной базе
    months_ = sorted({policy.date_end[5:7] for policy in PolicyBase.objects.all()})
    months = {}
    for month_ in months_:
        for i, month in months_rus.items():
            if int(month_) == int(i):
                months[i] = month

    if 'month_' in request.GET:
        month_ = request.GET['month_']
    else:
        month_ = datetime.datetime.now().month

    date_start = datetime.date(datetime.datetime.today().year, int(month_), 1)
    if month_ != 12:
        date_end = datetime.date(datetime.datetime.today().year, int(month_) + 1, 1)
    else:
        date_end = datetime.date(datetime.datetime.today().year + 1, 1, 1)
    policies = PolicyBase.objects.filter(date_end__range=(date_start, date_end))

    type_statistic = {}
    types = Type.objects.all()
    for type_ in types:
        # статистика по типам полисов
        policy_type = policies.filter(type=type_)
        if len(policy_type) > 0:
            temp_dict = {}
            count = len(policy_type)
            count_r = len(policy_type.filter(status__name='В работе'))
            count_p = len(policy_type.filter(status__name='Оформлен'))
            count_d = count - count_p - count_r
            temp_dict['count'] = count
            temp_dict['count_r'] = count_r
            temp_dict['percent_count_r'] = f'{format(float((count_r / count) * 100), ".1f")}%'
            temp_dict['count_p'] = count_p
            temp_dict['percent_count_p'] = f'{format(float((count_p / count) * 100), ".1f")}%'
            temp_dict['count_d'] = count_d
            temp_dict['percent_count_d'] = f'{format(float((count_d / count) * 100), ".1f")}%'
            type_statistic[type_] = temp_dict

    status_statistic = {}
    status_all = Status.objects.all()
    for status in status_all:
        # Статистика по отказам
        if status.name == 'Оформлен' or status.name == 'В работе':
            continue
        policy_status = policies.filter(status=status)
        if len(policy_status) > 0:
            status_statistic[status] = len(policy_status)

    if 'search' in request.GET:
        policies_pagi = policies.filter(
            Q(client__last_name__iregex=request.GET.get('search')) |
            Q(client__first_name__iregex=request.GET.get('search')) |
            Q(client__middle_name__iregex=request.GET.get('search')) |
            Q(bso__iregex=request.GET.get('search'))
        ).order_by('date_end')
    else:
        policies_pagi = policies.order_by('date_end')

    # ссылка с параметрами для пагинации
    link = '?'
    for key, value in request.GET.items():
        if key == 'page':
            continue
        link = link + f'{key}={value}&'

    paginator = Paginator(policies_pagi, 15)
    current_page = request.GET.get('page', 1)
    page = paginator.get_page(current_page)

    context = {
        'months': months,
        'month_': month_,
        'link': link,
        'page': page,
        'policies_pagi': page.object_list,
        'policies': policies,
        'paginator': paginator,
        'type_statistic': type_statistic,
        'status_statistic': status_statistic,
    }

    return render(request, 'prolongations/new_statistics.html', context)
