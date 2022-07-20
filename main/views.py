import datetime
import csv
import os

import openpyxl
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from openpyxl import load_workbook

from .models import Policy, Client, Channel, Company, User, Type, MortgagePolicy, Bank, Commission, Expenses, SaleReport
from .forms import UploadFileForm


# Месяцы
months_rus = {
    '01': 'Январь',
    '02': 'Февраль',
    '03': 'Март',
    '04': 'Апрель',
    '05': 'Май',
    '06': 'Июнь',
    '07': 'Июль',
    '08': 'Август',
    '09': 'Сентябрь',
    '10': 'Октябрь',
    '11': 'Ноябрь',
    '12': 'Декабрь',
}


def redirect_login(request):
    # Перенаправление с главной страницы
    return redirect('add_policy')


def get_start_end_date():
    # Получение первого и последнего дня текущего месяца
    now = datetime.datetime.now()
    date_start = datetime.datetime(now.year, now.month, 1).date()
    next_month = date_start + datetime.timedelta(35)
    date_end = datetime.datetime(next_month.year, next_month.month, 1).date()
    return date_start, date_end


@login_required(login_url='login')
def unload_files(request):
    # Выгрузка сверок в файл
    if request.user.admin:
        result = Policy.objects.filter(
            date_registration__lt=request.GET.get('date_end'),
            date_registration__gte=request.GET.get('date_start'),
            accept=False,
            sale_report=None,
        )
        if request.GET.get('Менеджер') != 'all':
            result = result.filter(user=request.GET.get('Менеджер'))
        if request.GET.get('Канал продаж') != 'all':
            result = result.filter(channel=request.GET.get('Канал продаж'))
        if request.GET.get('Страховая компания') != 'all':
            result = result.filter(company=request.GET.get('Страховая компания'))
        if request.GET.get('Тип полиса') != 'all':
            result = result.filter(type=request.GET.get('Тип полиса'))

    wb = openpyxl.Workbook()
    sheet = wb['Sheet']

    sheet['A1'] = '№'
    sheet['B1'] = 'Клиент'
    sheet['C1'] = 'Страховая компания'
    sheet['D1'] = 'Канал продаж'
    sheet['E1'] = 'Тип полиса'
    sheet['F1'] = 'Серия'
    sheet['G1'] = 'Номер'
    sheet['H1'] = 'Банк'
    sheet['I1'] = 'Премия'
    sheet['J1'] = 'КВ'
    sheet['K1'] = 'Дата оформления'
    sheet['L1'] = 'Дата начала действия'
    sheet['M1'] = 'Дата окончания действия'

    wb.save('main/file/mortgage.xlsx')

    wb = openpyxl.load_workbook('main/file/mortgage.xlsx')
    sheet = wb['Sheet']

    str_number = 2
    for policy in result:
        sheet[str_number][0].value = str_number - 1
        sheet[str_number][1].value = f'{policy.client.last_name} ' \
                                     f'{policy.client.first_name} ' \
                                     f'{policy.client.middle_name}'
        sheet[str_number][2].value = policy.company.name
        sheet[str_number][3].value = policy.channel.name
        sheet[str_number][4].value = policy.type.name
        sheet[str_number][5].value = policy.series
        sheet[str_number][6].value = policy.number
        sheet[str_number][7].value = policy.bank
        sheet[str_number][8].value = policy.sp
        sheet[str_number][9].value = policy.commission
        sheet[str_number][10].value = policy.date_registration
        sheet[str_number][11].value = policy.date_start
        sheet[str_number][12].value = policy.date_end

        str_number += 1

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="receivable.xlsx"'

    wb.save(response)

    return response


@login_required(login_url='login')
def unload_accept(request):
    # Выгрузка акцептованных полисов в xlsx
    if request.user.admin:
        result = Policy.objects.filter(
            date_registration__lt=request.GET.get('date_end'),
            date_registration__gte=request.GET.get('date_start'),
            accept=True,
            sale_report=None,
        )
        if request.GET.get('Менеджер') != 'all':
            result = result.filter(user=request.GET.get('Менеджер'))
        if request.GET.get('Канал продаж') != 'all':
            result = result.filter(channel=request.GET.get('Канал продаж'))
        if request.GET.get('Страховая компания') != 'all':
            result = result.filter(company=request.GET.get('Страховая компания'))
        if request.GET.get('Тип полиса') != 'all':
            result = result.filter(type=request.GET.get('Тип полиса'))

        wb = openpyxl.Workbook()
        sheet = wb['Sheet']

        sheet['A1'] = '№'
        sheet['B1'] = 'Клиент'
        sheet['C1'] = 'Страховая компания'
        sheet['D1'] = 'Канал продаж'
        sheet['E1'] = 'Тип полиса'
        sheet['F1'] = 'Серия'
        sheet['G1'] = 'Номер'
        sheet['H1'] = 'Банк'
        sheet['I1'] = 'Премия'
        sheet['J1'] = 'КВ'
        sheet['K1'] = 'Дата оформления'
        sheet['L1'] = 'Дата начала действия'
        sheet['M1'] = 'Дата окончания действия'

        wb.save('main/file/mortgage.xlsx')

        wb = openpyxl.load_workbook('main/file/mortgage.xlsx')
        sheet = wb['Sheet']

        str_number = 2
        for policy in result:
            sheet[str_number][0].value = str_number - 1
            sheet[str_number][1].value = f'{policy.client.last_name} ' \
                                         f'{policy.client.first_name} ' \
                                         f'{policy.client.middle_name}'
            sheet[str_number][2].value = policy.company.name
            sheet[str_number][3].value = policy.channel.name
            sheet[str_number][4].value = policy.type.name
            sheet[str_number][5].value = policy.series
            sheet[str_number][6].value = policy.number
            sheet[str_number][7].value = policy.bank
            sheet[str_number][8].value = policy.sp
            sheet[str_number][9].value = policy.commission
            sheet[str_number][10].value = policy.date_registration
            sheet[str_number][11].value = policy.date_start
            sheet[str_number][12].value = policy.date_end

            str_number += 1

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="receivable.xlsx"'

        wb.save(response)

        return response


@login_required(login_url='login')
def statistic(request):
    # просмотр статистики
    months_ = sorted({policy.date_registration.month for policy in Policy.objects.all()})
    months = {}
    for month_ in months_:
        for i, month in months_rus.items():
            if int(month_) == int(i):
                months[i] = month
    if 'month_' in request.GET:
        month_ = request.GET['month_']
    else:
        month_ = datetime.datetime.now().month

    date_start = datetime.date(datetime.datetime.today().year, int(month_), 1)
    if month_ != 12:
        date_end = datetime.date(datetime.datetime.today().year, int(month_) + 1, 1)
    else:
        date_end = datetime.date(datetime.datetime.today().year + 1, 1, 1)

    policies = Policy.objects.filter(date_registration__range=(date_start, date_end))

    all_statistic = {}
    user_statistic = {}
    all_sp = 0
    if request.user.admin:
        police_for_type = policies

        users = {policy.user: policy.user.id for policy in policies}
        for user, user_id in users.items():
            temp_dict = {}
            policy_user = police_for_type.filter(user=user_id)
            temp_dict['sp'] = int(sum(policy.sp for policy in policy_user))
            temp_dict['count'] = len(policy_user)
            user_statistic[user] = temp_dict
    else:
        police_for_type = policies.filter(user=request.user.id)

    types = {policy.type: policy.type.id for policy in police_for_type}
    for type, type_id in types.items():
        temp_dict = {}
        policy_type = police_for_type.filter(type=type_id)
        temp_dict['count'] = len(policy_type)
        temp_dict['sum_sp'] = int(sum(policy.sp for policy in policy_type))
        temp_dict['newbiz'] = len(policy_type.filter(status='newbiz'))
        temp_dict['prolongation'] = len(policy_type.filter(status='prolongation'))
        temp_dict['transition'] = len(policy_type.filter(status='transition'))
        temp_dict['payment'] = len(policy_type.filter(status='payment'))
        temp_dict['addendum'] = len(policy_type.filter(status='addendum'))
        all_statistic[type] = temp_dict
        all_sp += sum(policy.sp for policy in policy_type)

    context = {
        'all_statistic': all_statistic,
        'user_statistic': user_statistic,
        'all_sp': all_sp,
        'months': months,
    }
    return render(request, 'main/statistic.html', context)


@login_required(login_url='login')
def a_reporting(request):
    # Просмотр и выгрузка сверок
    if request.method == 'POST':
        if 'policy_id_for_delete' in request.POST and request.user.admin:
            Policy.objects.get(id=request.POST.get('policy_id_for_delete')).delete()

        if 'id_policy_for_accept' in request.POST:
            policy_for_accept = Policy.objects.get(id=request.POST.get('id_policy_for_accept'))
            policy_for_accept.accept = True
            policy_for_accept.save()

        if 'id_policy_for_edit' in request.POST:
            policy_for_edit = Policy.objects.get(id=request.POST.get('id_policy_for_edit'))

            full_name = request.POST['full_name_client']
            if f'{policy_for_edit.client.last_name} {policy_for_edit.client.first_name} ' \
               f'{policy_for_edit.client.middle_name}' != full_name:

                client_for_edit = policy_for_edit.client

                while full_name[-1] == ' ':
                    full_name = full_name[:-1]
                last_name = full_name.split()[0].capitalize()
                first_name = full_name.split()[1].capitalize()
                middle_name = ''
                if len(full_name.split()) > 2:
                    for name in full_name.split()[2:]:
                        middle_name = middle_name + name.capitalize() + " "
                    middle_name = middle_name[:-1]

                client_for_edit.last_name = last_name
                client_for_edit.first_name = first_name
                client_for_edit.middle_name = middle_name
                client_for_edit.save()

            policy_for_edit.type = Type.objects.get(id=request.POST.get('type'))
            policy_for_edit.status = request.POST.get('Тип_продажи')
            policy_for_edit.number = request.POST.get('number')
            policy_for_edit.series = request.POST.get('series')
            policy_for_edit.company = Company.objects.get(id=request.POST.get('company'))
            policy_for_edit.channel = Channel.objects.get(id=request.POST.get('channel'))
            policy_for_edit.sp = float(request.POST.get('sp').replace(',', '.'))
            policy_for_edit.commission = float(request.POST.get('commission').replace(',', '.'))
            policy_for_edit.date_registration = request.POST.get('date_registration')
            policy_for_edit.date_start = request.POST.get('date_start')
            policy_for_edit.date_end = request.POST.get('date_end')
            policy_for_edit.user = User.objects.get(id=request.POST.get('user'))
            policy_for_edit.save()

    selected = {}
    policy_list = []
    if request.GET.get('date_start') == None and request.GET.get('date_end') == None:
        date_start, date_end = get_start_end_date()
        date_start = date_start.strftime("%Y-%m-%d")
        date_end = date_end.strftime("%Y-%m-%d")
    else:
        date_start = request.GET.get('date_start')
        date_end = request.GET.get('date_end')
    if request.user.admin:
        users = User.objects.filter(admin=False, agent=False)
        result = Policy.objects.filter(accept=False, sale_report=None)
    else:
        users = User.objects.filter(id=request.user.id)
        result = Policy.objects.filter(user_id=request.user.id, accept=False, sale_report=None)
    company = Company.objects.all()
    channel = Channel.objects.all()
    type = Type.objects.all()

    if 'search' in request.GET:
        # поиск по страхователю или по номеру полиса
        selected['search'] = request.GET.get('search')
        result = result.filter(
            Q(client__last_name__iregex=request.GET.get('search')) |
            Q(client__first_name__iregex=request.GET.get('search')) |
            Q(client__middle_name__iregex=request.GET.get('search')) |
            Q(number__iregex=request.GET.get('search')) |
            Q(series__iregex=request.GET.get('search'))
        )
    else:
        result = result.filter(
            date_registration__lt=date_end,
            date_registration__gte=date_start
        )
        if 'Менеджер' in request.GET:
            if request.GET.get('Менеджер') != 'all':
                selected['manager'] = int(request.GET.get('Менеджер'))
                result = result.filter(user=request.GET.get('Менеджер'))
            if request.GET.get('Канал продаж') != 'all':
                selected['channel'] = int(request.GET.get('Канал продаж'))
                result = result.filter(channel=request.GET.get('Канал продаж'))
            if request.GET.get('Страховая компания') != 'all':
                selected['company'] = int(request.GET.get('Страховая компания'))
                result = result.filter(company=request.GET.get('Страховая компания'))
            if request.GET.get('Тип полиса') != 'all':
                selected['type'] = int(request.GET.get('Тип полиса'))
                result = result.filter(type=request.GET.get('Тип полиса'))

    if len(result) > 0:
        for policy in result:
            commission_rur = '{:.2f}'.format(policy.commission / 100 * policy.sp)
            temp_dict = {
                'status': policy.status,
                'type': policy.type,
                'series': policy.series,
                'number': policy.number,
                'company': policy.company,
                'channel': policy.channel,
                'sp': policy.sp,
                'commission': policy.commission,
                'commission_rur': commission_rur,
                'client': policy.client,
                'user': policy.user,
                'date_registration': policy.date_registration,
                'date_registration_for_edit': policy.date_registration.strftime("%Y-%m-%d"),
                'date_start': policy.date_start,
                'date_start_for_edit': policy.date_start.strftime("%Y-%m-%d"),
                'date_end': policy.date_end,
                'date_end_for_edit': policy.date_end.strftime("%Y-%m-%d"),
                'id': policy.id,
            }
            policy_list.append(temp_dict)

    # ссылка с параметрами для пагинации
    link = '?'
    for key, value in request.GET.items():
        if key == 'page':
            continue
        link = link + f'{key}={value}&'

    paginator = Paginator(policy_list, 15)
    current_page = request.GET.get('page', 1)
    page = paginator.get_page(current_page)

    data = {
        'users': users,
        'companies': company,
        'channels': channel,
        'types': type,
        'policy_list': page.object_list,
        'date_start': date_start,
        'date_end': date_end,
        'selected': selected,
        'page': page,
        'link': link,
        'paginator': paginator,
    }

    return render(request, 'main/reporting.html', data)


@login_required(login_url='login')
def addpolicy(request):
    # Добавление нового полиса
    error = ''
    errors_upload = {}
    text = ''
    text_upload = ''
    type = Type.objects.all()
    banks = Bank.objects.all()
    company = Company.objects.all()
    channel = Channel.objects.all()

    if request.method == 'POST':
        if 'upload' in request.POST:
            # загрузка продаж из файла
            wb = load_workbook(filename=request.FILES['file'])
            sheet = wb.worksheets[0]
            count = 0
            for row in range(2, sheet.max_row+1):
                # пропускаем пустую строчку
                if sheet[row][0].value is None:
                    continue

                status = '-'

                if sheet[row][0].value.lower() == 'новый бизнес':
                    status = 'newbiz'
                elif sheet[row][0].value.lower() == 'пролонгация':
                    status = 'prolongation'
                elif sheet[row][0].value.lower() == 'переход':
                    status = 'transition'
                elif sheet[row][0].value.lower() == 'аддендум':
                    status = 'addendum'

                if isinstance(sheet[row][7].value, str):
                    date_registration = f'{sheet[row][7].value[6:10]}-' \
                                        f'{sheet[row][7].value[3:5]}-' \
                                        f'{sheet[row][7].value[0:2]}'
                else:
                    date_registration = f'{sheet[row][7].value.year}-' \
                                        f'{sheet[row][7].value.month}-' \
                                        f'{sheet[row][7].value.day}'

                if isinstance(sheet[row][8].value, str):
                    date_start = f'{sheet[row][8].value[6:10]}-' \
                                        f'{sheet[row][8].value[3:5]}-' \
                                        f'{sheet[row][8].value[0:2]}'
                else:
                    date_start = f'{sheet[row][8].value.year}-' \
                                        f'{sheet[row][8].value.month}-' \
                                        f'{sheet[row][8].value.day}'

                if isinstance(sheet[row][9].value, str):
                    date_end = f'{sheet[row][9].value[6:10]}-' \
                                        f'{sheet[row][9].value[3:5]}-' \
                                        f'{sheet[row][9].value[0:2]}'
                else:
                    date_end = f'{sheet[row][9].value.year}-' \
                                        f'{sheet[row][9].value.month}-' \
                                        f'{sheet[row][9].value.day}'

                type_for_created, created = Type.objects.get_or_create(name=sheet[row][1].value)
                company_for_created, created = Company.objects.get_or_create(name=sheet[row][3].value)
                channel_for_created, created = Channel.objects.get_or_create(name=sheet[row][4].value)

                user = User.objects.get(
                    last_name=sheet[row][11].value.split()[0],
                    first_name=sheet[row][11].value.split()[1],
                )

                middle_name = ''
                if len(sheet[row][10].value.split()) == 3:
                    middle_name = sheet[row][10].value.split(" ")[2]
                elif len(sheet[row][10].value.split()) > 3:
                    middle_name = f'{sheet[row][10].value.split()[2]} {sheet[row][10].value.split()[3]}'

                client, created = Client.objects.get_or_create(
                    last_name=sheet[row][10].value.split()[0],
                    first_name=sheet[row][10].value.split()[1],
                    middle_name=middle_name,
                )

                if len(str(sheet[row][2].value).split()) == 2:
                    series = sheet[row][2].value.split()[0]
                    number = sheet[row][2].value.split()[1]
                else:
                    series = ''
                    number = sheet[row][2].value

                policy, created = Policy.objects.get_or_create(
                    series=series,
                    number=number,
                    defaults={
                        'status': status,
                        'type': type_for_created,
                        'company': company_for_created,
                        'channel': channel_for_created,
                        'sp': float(str(sheet[row][5].value).replace(' ', '').replace(',', '.')),
                        'commission': float(str(sheet[row][6].value).replace(' ', '').replace(',', '.')),
                        'date_registration': date_registration,
                        'date_start': date_start,
                        'date_end': date_end,
                        'user': user,
                        'client': client,
                    }
                )

                if created:
                    count += 1
                else:
                    errors_upload[f'{series} {number}'] = 'Полис уже есть в базе'

            text_upload = f'Загруженно {count} из {count + len(errors_upload)}'

            if len(errors_upload) > 0:
                wb_errors = openpyxl.Workbook()
                sheet = wb_errors['Sheet']

                sheet['A1'] = 'Номер'
                sheet['B1'] = 'БСО'
                sheet['C1'] = 'Ошибка'

                wb_errors.save(f'main/file/wb_errors.xlsx')

                wb_errors = openpyxl.load_workbook(f'main/file/wb_errors.xlsx')
                sheet = wb_errors['Sheet']

                str_number = 2
                count_policy = 1
                for policy, error in errors_upload.items():
                    sheet[str_number][0].value = count_policy
                    sheet[str_number][1].value = policy
                    sheet[str_number][2].value = error

                    str_number += 1

                response = HttpResponse(content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename="wb_errors.xlsx"'

                wb_errors.save(response)

                return response

        else:
            full_name = request.POST['full_name']
            while full_name[-1] == ' ':
                full_name = full_name[:-1]
            last_name = full_name.split()[0].capitalize()
            first_name = full_name.split()[1].capitalize()
            middle_name = ''
            if len(full_name.split()) > 2:
                for name in full_name.split()[2:]:
                    middle_name = middle_name + name.capitalize() + " "
                middle_name = middle_name[:-1]

            client, created = Client.objects.get_or_create(
                first_name=first_name,
                middle_name=middle_name,
                last_name=last_name,
                birthday=f"{request.POST.get('birthday')[6:10]}-{request.POST.get('birthday')[3:5]}-"
                         f"{request.POST.get('birthday')[0:2]}",
                defaults={
                    'phone': request.POST.get('phone'),
                    'email': request.POST.get('email')
                }
            )

            if request.POST.get('Тип продажи') == 'payment':
                # очередной взнос
                policy = Policy(
                    number=request.POST.get('number'),
                    series=request.POST.get('series'),
                    company_id=request.POST.get('Страховая_компания'),
                    user_id=request.user.id,
                    client_id=client.id,
                    channel_id=request.POST.get('Канал_продаж'),
                    type_id=request.POST.get('Тип_полиса'),
                    date_registration=request.POST.get('date_registration'),
                    date_start=request.POST.get('date_start'),
                    date_end=request.POST.get('date_end'),
                    commission=float(request.POST.get('commission').replace(',', '.')),
                    sp=float(request.POST.get('sp').replace(',', '.')),
                    status=request.POST.get('Тип продажи'),
                )
                policy.save()

                text = 'Запись успешно добавлена'
                if Type.objects.get(id=request.POST.get('Тип_полиса')).name == 'Ипотечный':
                    policy.bank = request.POST.get('bank')
                    policy.save()
                if request.POST.get('Оплата') == 'cash':
                    policy.type_pay = True
                    policy.save()
                if request.POST.get('credit') == 'credit':
                    policy.credit = True
                    policy.save()

                commission_objects = Commission.objects.filter(
                    type=policy.type,
                    channel=policy.channel,
                    company=policy.company,
                    date_start__lte=policy.date_registration,
                )

                if len(commission_objects) > 0:
                    if Type.objects.get(id=request.POST.get('Тип_полиса')).name == 'Ипотечный':
                        commission_objects = commission_objects.filter(bank=Bank.objects.get(name=policy.bank))

                    if len(commission_objects) > 0:
                        policy.commission = commission_objects.order_by('-date_start')[0].value
                        policy.save()

            else:
                # добавление нового полиса
                policy, created = Policy.objects.get_or_create(
                    number=request.POST.get('number'),
                    series=request.POST.get('series'),
                    company_id=request.POST.get('Страховая_компания'),
                    defaults={
                        'user_id': request.user.id,
                        'client_id': client.id,
                        'channel_id': request.POST.get('Канал_продаж'),
                        'type_id': request.POST.get('Тип_полиса'),
                        'date_registration': request.POST.get('date_registration'),
                        'date_start': request.POST.get('date_start'),
                        'date_end': request.POST.get('date_end'),
                        'commission': float(request.POST.get('commission').replace(',', '.')),
                        'sp': float(request.POST.get('sp').replace(',', '.')),
                        'status': request.POST.get('Тип продажи'),
                        }
                    )

                if not created:
                    error = f'Полис с номером {request.POST.get("series")} {request.POST.get("number")} уже есть в базе'

                else:
                    text = 'Запись успешно добавлена'
                    if Type.objects.get(id=request.POST.get('Тип_полиса')).name == 'Ипотечный':
                        policy.bank = request.POST.get('bank')
                        policy.save()
                    if request.POST.get('Оплата') == 'cash':
                        policy.type_pay = True
                        policy.save()
                    if request.POST.get('credit') == 'credit':
                        policy.credit = True
                        policy.save()

                    commission_objects = Commission.objects.filter(
                        type=policy.type,
                        channel=policy.channel,
                        company=policy.company,
                        date_start__lte=policy.date_registration,
                    )

                    if len(commission_objects) > 0:
                        if Type.objects.get(id=request.POST.get('Тип_полиса')).name == 'Ипотечный':
                            commission_objects = commission_objects.filter(bank=Bank.objects.get(name=policy.bank))

                        if len(commission_objects) > 0:
                            policy.commission = commission_objects.order_by('-date_start')[0].value
                            policy.save()

    data = {
        'types': type,
        'companys': company,
        'channels': channel,
        'banks': banks,
        'error': error,
        'text': text,
        'text_upload': text_upload,
    }

    return render(request, 'main/osago.html', data)


@login_required(login_url='login')
def upload_policy(request):
    form = UploadFileForm()
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            text = 'Загружено'

            with open('main/file/date.csv', 'wb') as file:
                for row in request.FILES['file'].chunks():
                    file.write(row)

            with open('main/file/date.csv', 'r', encoding='cp1251', newline='') as file_1:
                reader = csv.DictReader(file_1, delimiter=';')
                for row in reader:
                    pass

    context = {}

    return render(request, 'main/upload_policy.html', context)


@login_required(login_url='login')
def register_user(request):
    # Регистрация нового пользователя
    if request.user.admin:
        error = ''
        if request.method == 'POST':
            if len(User.objects.filter(username=request.POST.get('username'))) > 0:
                error = 'Этот логин уже используется'
            else:
                user = User.objects.create_user(
                    username=request.POST.get('username'),
                    password=request.POST.get('password'),
                    last_name=request.POST.get('last_name'),
                    first_name=request.POST.get('first_name'),
                    middle_name='',
                    agent=False,
                )
                if request.POST.get('middle_name'):
                    user.middle_name = request.POST.get('middle_name')
                    user.save()

        return render(request, 'registration/registration.html', context={'error': error})

    else:
        return render(request, 'registration/registration.html',
                      context={'error': 'Для регистрации обратитесь к администратору!'})


@login_required(login_url='login')
def policy_edit(request):
    text = ''
    if request.user.admin:
        policy = Policy.objects.get(id=request.GET.get('id'))
        channel = Channel.objects.all()
        company = Company.objects.all()
        date_start = policy.date_start.strftime("%Y-%m-%d")
        date_end = policy.date_end.strftime("%Y-%m-%d")
        sp = str(policy.sp).replace(',', '.')

        if request.method == 'POST':
            policy = Policy.objects.get(id=request.POST.get('id'))
            policy.number = request.POST.get('number')
            policy.series = request.POST.get('series')
            policy.company = Company.objects.get(id=request.POST.get('company'))
            policy.channel = Channel.objects.get(id=request.POST.get('channel'))
            policy.sp = float(request.POST.get('sp').replace(',', '.'))
            policy.date_start = request.POST.get('date_start')
            policy.date_end = request.POST.get('date_end')
            policy.save()

            sp = request.POST.get('sp')
            text = 'Сохранено'

        context = {
            'policy': policy,
            'text': text,
            'channel': channel,
            'company': company,
            'date_start': date_start,
            'date_end': date_end,
            'sp': sp,
        }

        return render(request, 'main/policy_edit_div.html', context)


@login_required(login_url='login')
def unload_mortgage(request):
    # Выгрузка статистики в файл xlsx
    if request.user.admin:
        date_end = datetime.datetime.strptime(f'{request.GET.get("year")}-{request.GET.get("month")}-01', '%Y-%m-%d')
        result = MortgagePolicy.objects.filter(date_end=date_end, date_at__gte=request.GET.get('date_at'))
        if request.GET.get('bank') != 'all':
            result = result.filter(bank_id=request.GET.get('bank'))

        wb = openpyxl.Workbook()
        sheet = wb['Sheet']

        sheet['A1'] = 'Банк'
        sheet['B1'] = 'Тип полиса'
        sheet['C1'] = 'Дата окончания'
        sheet['D1'] = 'Клиент'
        sheet['E1'] = 'Дата рождения'
        sheet['F1'] = 'Телефон'
        sheet['G1'] = 'Почта'

        wb.save('main/file/mortgage.xlsx')

        wb = openpyxl.load_workbook('main/file/mortgage.xlsx')
        sheet = wb['Sheet']

        str_number = 2
        for policy in result:
            sheet[str_number][0].value = policy.bank.name
            if policy.type_mortgage == 'all':
                sheet[str_number][1].value = 'Все риски'
            elif policy.type_mortgage == 'property':
                sheet[str_number][1].value = 'Имущество'
            elif policy.type_mortgage == 'life':
                sheet[str_number][1].value = 'Жизнь'
            sheet[str_number][2].value = policy.date_end
            sheet[str_number][3].value = f'{policy.client.last_name} ' \
                                         f'{policy.client.first_name} ' \
                                         f'{policy.client.middle_name}'
            sheet[str_number][4].value = policy.client.birthday
            sheet[str_number][5].value = policy.client.phone
            sheet[str_number][6].value = policy.client.email

            str_number += 1

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="mortgage.xlsx"'

        wb.save(response)

        return response


@login_required(login_url='login')
def upload_mortgage(request):
    # Загрузка заявок ипотечных заявок
    text = ''
    step_upload = 0
    if request.method == 'POST':
        # первый этап загрузки файла
        if 'upload' in request.POST:
            step_upload = 1
            headers_for_upload = {
                'ФИО': 'full_name',
                'Дата рождения': 'birthday',
                'Телефон': 'phone',
                'Почта': 'email',
                'Банк': 'bank',
                'Дата окончания': 'date_end',
            }

            wb = openpyxl.load_workbook(filename=request.FILES['file'])
            sheet = wb.worksheets[0]

            col = 0
            headers = {}
            while True:
                try:
                    if sheet[1][col].value is None or sheet[1][col].value == '':
                        break
                    headers[col] = sheet[1][col].value
                    col += 1
                except:
                    break

            wb.save('main/file/upload.xlsx')

            context = {
                'step_upload': step_upload,
                'headers': headers,
                'headers_for_upload': headers_for_upload,
            }

            return render(request, 'main/upload_mortgage.html', context)

        # второй этап загрузки агентских продаж из файла
        if 'step_upload' in request.POST:
            errors = {}

            full_name_col = int(request.POST['full_name'])
            birthday_col = int(request.POST['birthday'])
            phone_col = int(request.POST['phone'])
            email_col = int(request.POST['email'])
            bank_col = int(request.POST['bank'])
            date_end_col = int(request.POST['date_end'])

            wb = openpyxl.load_workbook('main/file/upload.xlsx')
            sheet = wb.worksheets[0]

            number = 0
            row = 2
            while True:
                if sheet[row][0].value is None or sheet[row][0].value == '':
                    break

                if not isinstance(sheet[row][birthday_col].value, datetime.date):
                    errors[sheet[row][full_name_col].value] = 'Дата рождения - нужен формат даты'
                    row += 1
                    continue
                if not isinstance(sheet[row][date_end_col].value, datetime.date):
                    errors[sheet[row][full_name_col].value] = 'Дата окончания - нужен формат даты'
                    row += 1
                    continue

                full_name = sheet[row][full_name_col].value

                while full_name[-1] == ' ':
                    full_name = full_name[:-1]
                last_name = full_name.split()[0].capitalize()
                first_name = full_name.split()[1].capitalize()
                middle_name = ''
                if len(full_name.split()) > 2:
                    for name in full_name.split()[2:]:
                        middle_name = middle_name + name.capitalize() + " "
                    middle_name = middle_name[:-1]

                client, created = Client.objects.get_or_create(
                    last_name=last_name,
                    first_name=first_name,
                    middle_name=middle_name,
                    birthday=sheet[row][birthday_col].value,
                    defaults={
                        'phone': sheet[row][phone_col].value,
                        'email': sheet[row][email_col].value,
                    }
                )

                bank, created = Bank.objects.get_or_create(name=sheet[row][bank_col].value)

                new_mortgage = MortgagePolicy(
                    user_id=request.user.id,
                    date_end=sheet[row][date_end_col].value,
                    client=client,
                    bank=bank,
                )
                new_mortgage.save()
                number += 1
                row += 1

            os.remove('main/file/upload.xlsx')

            # сохраняем файл с ошибками загрузки
            if len(errors) > 0:
                errors_unload = 1
                errors_count = len(errors)

                wb = openpyxl.Workbook()
                sheet = wb['Sheet']

                sheet['A1'] = '№'
                sheet['B1'] = 'ФИО'
                sheet['C1'] = 'Тип ошибки'

                str_number = 2
                for error, text in errors.items():
                    sheet[str_number][0].value = str_number - 1
                    sheet[str_number][1].value = error
                    sheet[str_number][2].value = text
                    str_number += 1

                wb.save('main/file/errors.xlsx')

                context = {
                    'errors_count': errors_count,
                    'errors_unload': errors_unload,
                    'step_upload': step_upload,
                    'number': number,
                }

                return render(request, 'main/upload_mortgage.html', context)

        # выгрузка ошибок при загрузке
        if 'errors_unload' in request.POST:
            wb = openpyxl.load_workbook('main/file/errors.xlsx')

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="receivable.xlsx"'

            wb.save(response)

            return response

    context = {
        'text': text,
        'step_upload': step_upload,
    }

    return render(request, 'main/upload_mortgage.html', context)


@login_required(login_url='login')
def mortgage(request):
    text = ''
    users_statistic = {}
    if request.method == 'POST':
        client, created = Client.objects.get_or_create(
            first_name=request.POST.get('first_name'),
            middle_name=request.POST.get('middle_name'),
            last_name=request.POST.get('last_name'),
            birthday=request.POST.get('birthday'),
            defaults={
                'phone': request.POST.get('phone'),
                'email': request.POST.get('email')
            }
        )

        date_end = datetime.datetime.strptime(f'{request.POST.get("year")}-{request.POST.get("month")}-01', '%Y-%m-%d')
        MortgagePolicy(bank_id=request.POST.get('bank'),
                       type_mortgage=request.POST['type_policy_mortgage'],
                       user_id=request.user.id,
                       client_id=client.id,
                       date_end=date_end).save()
        text = 'Полис добавлен'
    if request.user.admin:
        mortgage_policy = MortgagePolicy.objects.all()
        date_target = datetime.datetime.today().date()
        for i in range(1, 7):
            policyes = MortgagePolicy.objects.filter(date_at=date_target)
            if len(policyes) > 0:
                users_set = {policy.user for policy in policyes}
                temp_dict = {user: len(MortgagePolicy.objects.filter(user=user)) for user in users_set}
                users_statistic[date_target] = temp_dict
            date_target = date_target - datetime.timedelta(days=-1)
    else:
        mortgage_policy = MortgagePolicy.objects.filter(user_id=request.user.id)

    statistic_2022 = {}
    statistic_2022['all_count'] = len(mortgage_policy.filter(
        date_end__gte=datetime.datetime.strptime('2022-01-01', '%Y-%m-%d'),
        date_end__lt=datetime.datetime.strptime('2023-01-01', '%Y-%m-%d')))
    for i in range(1, 13):
        if i < 9:
            statistic_2022['0' + str(i)] = len(mortgage_policy.filter(
        date_end__gte=datetime.datetime.strptime(f'2022-{"0" + str(i)}-01', '%Y-%m-%d'),
        date_end__lt=datetime.datetime.strptime(f'2022-{"0" + str(i + 1)}-01', '%Y-%m-%d')))
        elif i == 9:
            statistic_2022['0' + str(i)] = len(mortgage_policy.filter(
                date_end__gte=datetime.datetime.strptime(f'2022-{"0" + str(i)}-01', '%Y-%m-%d'),
                date_end__lt=datetime.datetime.strptime(f'2022-{str(i + 1)}-01', '%Y-%m-%d')))
        elif i == 10:
            statistic_2022[str(i)] = len(mortgage_policy.filter(
                date_end__gte=datetime.datetime.strptime(f'2022-{str(i)}-01', '%Y-%m-%d'),
                date_end__lt=datetime.datetime.strptime(f'2022-{str(i + 1)}-01', '%Y-%m-%d')))
        elif i < 12:
            statistic_2022[str(i)] = len(mortgage_policy.filter(
                date_end__gte=datetime.datetime.strptime(f'2022-{str(i)}-01', '%Y-%m-%d'),
                date_end__lt=datetime.datetime.strptime(f'2022-{str(i + 1)}-01', '%Y-%m-%d')))
        else:
            statistic_2022[str(i)] = len(mortgage_policy.filter(
                date_end__gte=datetime.datetime.strptime(f'2022-{str(i)}-01', '%Y-%m-%d'),
                date_end__lt=datetime.datetime.strptime(f'2023-{"0" + str(i - 11)}-01', '%Y-%m-%d')))

    statistic_2023 = {}
    statistic_2023['all_count'] = len(mortgage_policy.filter(
        date_end__gte=datetime.datetime.strptime('2023-01-01', '%Y-%m-%d'),
        date_end__lt=datetime.datetime.strptime('2024-01-01', '%Y-%m-%d')))
    for i in range(1, 13):
        if i < 9:
            statistic_2023['0' + str(i)] = len(mortgage_policy.filter(
                date_end__gte=datetime.datetime.strptime(f'2023-{"0" + str(i)}-01', '%Y-%m-%d'),
                date_end__lt=datetime.datetime.strptime(f'2023-{"0" + str(i + 1)}-01', '%Y-%m-%d')))
        elif i == 9:
            statistic_2023['0' + str(i)] = len(mortgage_policy.filter(
                date_end__gte=datetime.datetime.strptime(f'2023-{"0" + str(i)}-01', '%Y-%m-%d'),
                date_end__lt=datetime.datetime.strptime(f'2023-{str(i + 1)}-01', '%Y-%m-%d')))
        elif i == 10:
            statistic_2023[str(i)] = len(mortgage_policy.filter(
                date_end__gte=datetime.datetime.strptime(f'2023-{str(i)}-01', '%Y-%m-%d'),
                date_end__lt=datetime.datetime.strptime(f'2023-{str(i + 1)}-01', '%Y-%m-%d')))
        elif i < 12:
            statistic_2023[str(i)] = len(mortgage_policy.filter(
                date_end__gte=datetime.datetime.strptime(f'2023-{str(i)}-01', '%Y-%m-%d'),
                date_end__lt=datetime.datetime.strptime(f'2023-{str(i + 1)}-01', '%Y-%m-%d')))
        else:
            statistic_2023[str(i)] = len(mortgage_policy.filter(
                date_end__gte=datetime.datetime.strptime(f'2023-{str(i)}-01', '%Y-%m-%d'),
                date_end__lt=datetime.datetime.strptime(f'2024-{"0" + str(i - 11)}-01', '%Y-%m-%d')))

    banks = Bank.objects.all()
    date_at = datetime.datetime.today().strftime("%Y-%m-%d")

    # для статистики по менеджерам
    date_now = datetime.datetime.now().date()
    date_start_for_statistic = date_now - datetime.timedelta(days=7)
    mortgage_for_statistic = MortgagePolicy.objects.filter(
        date_at__gt=date_start_for_statistic,
        date_at__lte=date_now,
    )
    user_ids = set(policy.user.id for policy in mortgage_for_statistic)
    mortgage_user_statistic = {}
    date_for_headers = [(date_now - datetime.timedelta(days=i)).strftime("%d.%m") for i in reversed(range(7))]
    for user_id in user_ids:
        mortgage_user = mortgage_for_statistic.filter(user_id=user_id)
        temp_dict = {}
        temp_dict['count'] = len(mortgage_user)
        for i in reversed(range(7)):
            temp_dict[(date_now - datetime.timedelta(days=i)).strftime("%d.%m")] = len(mortgage_user.filter(
                date_at=date_now - datetime.timedelta(days=i)))
        user_ = User.objects.get(id=user_id)
        user_name = f'{user_.last_name} {user_.first_name}'
        mortgage_user_statistic[user_name] = temp_dict

    context = {
        'banks': banks,
        'date_at': date_at,
        'mortgage_policy': mortgage_policy,
        'statistic_2023': statistic_2023,
        'statistic_2022': statistic_2022,
        'text': text,
        'users_statistic': users_statistic,
        'mortgage_user_statistic': mortgage_user_statistic,
        'date_for_headers': date_for_headers,
    }

    return render(request, 'main/Mortgage.html', context)


@login_required(login_url='login')
def commission(request):
    if request.method == 'POST':
        if Type.objects.get(id=request.POST.get('type')).name == 'Ипотечный':
            commission, created = Commission.objects.get_or_create(
                channel_id=request.POST.get('id'),
                company_id=request.POST.get('company'),
                type_id=request.POST.get('type'),
                bank_id=request.POST.get('bank'),
                date_start=request.POST.get('date_start'),
                defaults={
                    'value': float(request.POST.get('value').replace(',', '.')),
                }
            )
        else:
            commission, created = Commission.objects.get_or_create(
                channel_id=request.POST.get('id'),
                company_id=request.POST.get('company'),
                type_id=request.POST.get('type'),
                date_start=request.POST.get('date_start'),
                defaults={
                    'value': float(request.POST.get('value').replace(',', '.')),
                }
            )

        if created == False:
            commission.value = float(request.POST.get('value').replace(',', '.'))
            commission.save()

        policyes = Policy.objects.filter(date_registration__gte=commission.date_start,
                                         channel=commission.channel,
                                         company=commission.company,
                                         type=commission.type,
                                         accept=False)

        if len(policyes) > 0:
            if commission.type.name == 'Ипотечный':
                policyes = policyes.filter(bank=commission.bank)

                if len(policyes) > 0:
                    for policy in policyes:
                        policy.commission = (Commission.objects.filter(
                            type=commission.type,
                            channel=commission.channel,
                            company=commission.company,
                            date_start__lte=policy.date_registration,
                            bank=commission.bank)).order_by('-date_start')[0].value
                        policy.save()
            else:
                for policy in policyes:
                    policy.commission = (Commission.objects.filter(
                        type=commission.type,
                        date_start__lte=policy.date_registration,
                        channel=commission.channel,
                        company=commission.company,
                    )).order_by('-date_start')[0].value
                    policy.save()

    channel = Channel.objects.all()
    commission = Commission.objects.all()
    company = Company.objects.all()
    bank = Bank.objects.all()
    type = Type.objects.all()
    date_now = datetime.datetime.now().strftime("%Y-%m-%d")

    context = {
        'channel': channel,
        'company': company,
        'date_now': date_now,
        'commission': commission,
        'bank': bank,
        'type': type,
    }
    return render(request, 'main/commission.html', context)


@login_required(login_url='login')
def commission_delete(request):
    if request.user.admin:
        if request.method == 'POST':
            Commission.objects.get(id=request.POST.get('id')).delete()
            return HttpResponseRedirect('/commission')


@login_required(login_url='login')
def accept(request):
    # Просмотр и выгрузка проведенных полисов
    if request.method == 'POST' and request.user.admin and 'policy_id_for_return' in request.POST:
        policy_for_return = Policy.objects.get(id=request.POST.get('policy_id_for_return'))
        policy_for_return.accept = False
        policy_for_return.save()

    selected = {}
    policy_list = []
    if request.GET.get('date_start') == None and request.GET.get('date_end') == None:
        date_start, date_end = get_start_end_date()
        date_start = date_start.strftime("%Y-%m-%d")
        date_end = date_end.strftime("%Y-%m-%d")
    else:
        date_start = request.GET.get('date_start')
        date_end = request.GET.get('date_end')
    if request.user.admin:
        users = User.objects.all()
        result = Policy.objects.filter(accept=True, sale_report=None)
    else:
        users = User.objects.filter(id=request.user.id)
        result = Policy.objects.filter(user_id=request.user.id, accept=True, sale_report=None)
    company = Company.objects.all()
    channel = Channel.objects.all()
    type = Type.objects.all()

    if 'search' in request.GET:
        # поиск по страхователю или по номеру полиса
        selected['search'] = request.GET.get('search')
        result = result.filter(
            Q(client__last_name__iregex=request.GET.get('search')) |
            Q(client__first_name__iregex=request.GET.get('search')) |
            Q(client__middle_name__iregex=request.GET.get('search')) |
            Q(number__iregex=request.GET.get('search')) |
            Q(series__iregex=request.GET.get('search'))
        )
    else:
        result = result.filter(
            date_registration__lt=date_end,
            date_registration__gte=date_start
        )
        if 'Менеджер' in request.GET:
            if request.GET.get('Менеджер') != 'all':
                selected['manager'] = int(request.GET.get('Менеджер'))
                result = result.filter(user=request.GET.get('Менеджер'))
            if request.GET.get('Канал продаж') != 'all':
                selected['channel'] = int(request.GET.get('Канал продаж'))
                result = result.filter(channel=request.GET.get('Канал продаж'))
            if request.GET.get('Страховая компания') != 'all':
                selected['company'] = int(request.GET.get('Страховая компания'))
                result = result.filter(company=request.GET.get('Страховая компания'))
            if request.GET.get('Тип полиса') != 'all':
                selected['type'] = int(request.GET.get('Тип полиса'))
                result = result.filter(type=request.GET.get('Тип полиса'))

    if len(result) > 0:
        for policy in result:
            commission_rur = '{:.2f}'.format(policy.commission / 100 * policy.sp)
            temp_dict = {
                'status': policy.status,
                'type': policy.type,
                'series': policy.series,
                'number': policy.number,
                'company': policy.company,
                'channel': policy.channel,
                'sp': policy.sp,
                'commission': policy.commission,
                'commission_rur': commission_rur,
                'client': policy.client,
                'user': policy.user,
                'date_registration': policy.date_registration,
                'date_registration_for_edit': policy.date_registration.strftime("%Y-%m-%d"),
                'date_start': policy.date_start,
                'date_start_for_edit': policy.date_start.strftime("%Y-%m-%d"),
                'date_end': policy.date_end,
                'date_end_for_edit': policy.date_end.strftime("%Y-%m-%d"),
                'id': policy.id,
            }
            policy_list.append(temp_dict)

    # ссылка с параметрами для пагинации
    link = '?'
    for key, value in request.GET.items():
        if key == 'page':
            continue
        link = link + f'{key}={value}&'

    paginator = Paginator(policy_list, 15)
    current_page = request.GET.get('page', 1)
    page = paginator.get_page(current_page)

    data = {
        'users': users,
        'companies': company,
        'channels': channel,
        'types': type,
        'policy_list': page.object_list,
        'paginator': paginator,
        'page': page,
        'date_start': date_start,
        'date_end': date_end,
        'selected': selected,
        'link': link,
    }
    return render(request, 'main/accept.html', data)


@login_required(login_url='login')
def add_type_channel_company(request):
    # Добавление страховой компании, канала продаж и типа полиса
    text_company = ''
    text_channel = ''
    text_bank = ''
    text_type = ''

    if request.method == 'POST':
        if 'new_channel' in request.POST:
            new_channel, created = Channel.objects.get_or_create(name=request.POST['new_channel'])

            if created:
                text_channel = 'Новый канал продаж добавлен'
            else:
                text_channel = 'Такой канал продаж уже существует'

        if 'new_type' in request.POST:
            new_type, created = Type.objects.get_or_create(name=request.POST['new_type'])

            if created:
                text_type = 'Новый тип полиса добавлен'
            else:
                text_type = 'Такой тип полиса уже существует'

        if 'new_company' in request.POST:
            new_company, created = Company.objects.get_or_create(name=request.POST['new_company'])

            if created:
                text_company = 'Новая страховая компания добавлена'
            else:
                text_company = 'Такая страховая компания уже существует'

        if 'new_bank' in request.POST:
            new_bank, created = Bank.objects.get_or_create(name=request.POST['new_bank'])

            if created:
                text_bank = 'Новый банк добавлен'
            else:
                text_bank = 'Такой банк уже существует'

    channel = Channel.objects.all()
    type = Type.objects.all()
    company = Company.objects.all()
    bank = Bank.objects.all()

    context = {
        'text_company': text_company,
        'text_channel': text_channel,
        'text_bank': text_bank,
        'text_type': text_type,
        'channel': channel,
        'type': type,
        'bank': bank,
        'company': company,
    }

    return render(request, 'main/add_type_channel_company.html', context)


@login_required(login_url='login')
def get_expenses(request):
    # расходы и статистика по ним
    selected = {}
    expenses = Expenses.objects.filter(sale_report=None)
    date_start, date_end = get_start_end_date()
    date_start = date_start.strftime("%Y-%m-%d")
    date_end = date_end.strftime("%Y-%m-%d")

    if request.user.username == 'MLondorenko' or request.user.username == 'DMurovtsev':
        if request.method == 'POST':
            if 'expenses_for_delete' in request.POST:
                # удаление расхода
                Expenses.objects.get(id=request.POST['expenses_id_for_delete']).delete()

            if 'create_expenses' in request.POST:
                # добавление расхода
                expenses_new, created = Expenses.objects.get_or_create(
                    name=request.POST['name_expenses'],
                    date_expenses=request.POST['date_expenses'],
                    defaults={
                        'value': float(request.POST['value_expenses'].replace(',', '.')),
                    }
                )
                if not created:
                    old_value = expenses_new.value
                    new_value = float(request.POST['value_expenses'].replace(',', '.')) + float(old_value)
                    expenses_new.value = new_value
                    expenses_new.save()

                if request.POST['salary'] == 'yes':
                    expenses_new.salary = True
                    expenses_new.save()

        if 'filter_expenses' in request.GET:
            # фильтрация расходов
            selected['date_start'] = request.GET['date_start']
            selected['date_end'] = request.GET['date_end']
            expenses_for_filter = Expenses.objects.filter(
                date_expenses__gte=request.GET['date_start'],
                date_expenses__lte=request.GET['date_end'],
            )
            if request.GET['name_expenses'] != 'all':
                expenses_for_filter = expenses_for_filter.filter(name=request.GET['name_expenses'])
                selected['name_expenses'] = request.GET['name_expenses']
            if request.GET['salary'] != 'all':
                selected['salary'] = request.GET['salary']
                if request.GET['salary'] == 'yes':
                    expenses_for_filter = expenses_for_filter.filter(salary=True)
                else:
                    expenses_for_filter = expenses_for_filter.filter(salary=False)
        else:
            expenses_for_filter = expenses

        sum_expenses = sum(ex.value for ex in expenses_for_filter)
        sum_for_final_statistic = {}
        for ex in expenses_for_filter:
            if ex.name in sum_for_final_statistic:
                sum_for_final_statistic[ex.name] += ex.value
            else:
                sum_for_final_statistic[ex.name] = ex.value

        context = {
            'expenses': expenses,
            'expenses_for_filter': expenses_for_filter,
            'date_start': date_start,
            'date_end': date_end,
            'selected': selected,
            'sum_expenses': sum_expenses,
            'sum_for_final_statistic': sum_for_final_statistic,
        }

        return render(request, 'main/expenses.html', context)


@login_required(login_url='login')
def unload_expenses(request):
    # выгрзука расходов в xlsx
    if request.user.username == 'MLondorenko' or request.user.username == 'DMurovtsev':
        if request.method == 'POST':
            if 'filter_expenses' in request.POST:
                # выгрузка расходов
                expenses_for_filter = Expenses.objects.filter(
                    date_expenses__gte=request.POST['date_start'],
                    date_expenses__lte=request.POST['date_end'],
                    sale_report=None,
                )
                if request.POST['name_expenses'] != 'all':
                    expenses_for_filter = expenses_for_filter.filter(name=request.POST['name_expenses'])
                if request.POST['salary'] != 'all':
                    if request.POST['salary'] == 'yes':
                        expenses_for_filter = expenses_for_filter.filter(salary=True)
                    else:
                        expenses_for_filter = expenses_for_filter.filter(salary=False)

                wb = openpyxl.Workbook()
                sheet = wb['Sheet']

                sheet['A1'] = '№'
                sheet['B1'] = 'Наименование'
                sheet['C1'] = 'Дата расхода'
                sheet['D1'] = 'Сумма'

                wb.save('main/file/expenses.xlsx')

                wb = openpyxl.load_workbook('main/file/expenses.xlsx')
                sheet = wb['Sheet']

                str_number = 2
                count = 1
                for expenses in expenses_for_filter:
                    sheet[str_number][0].value = count
                    sheet[str_number][1].value = expenses.name
                    sheet[str_number][2].value = expenses.date_expenses
                    sheet[str_number][3].value = expenses.value

                    str_number += 1
                    count += 1

                response = HttpResponse(content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename="expenses.xlsx"'

                wb.save(response)

                return response


@login_required(login_url='login')
def create_sale_report(request):
    # создание акта продаж
    if request.user.admin and request.method == 'POST':
        new_sale_report = SaleReport(name=request.POST['name_sale_report'])
        new_sale_report.save()

        policies = Policy.objects.filter(
           date_registration__gte=request.POST['date_start_for_report'],
           date_registration__lte=request.POST['date_end_for_report'],
           accept=True,
           sale_report=None,
        )

        expenses = Expenses.objects.filter(
           date_expenses__gte=request.POST['date_start_for_report'],
           date_expenses__lte=request.POST['date_end_for_report'],
           sale_report=None,
        )

        for policy in policies:
            policy.sale_report = new_sale_report
            policy.save()

        for expenses in expenses:
            expenses.sale_report = new_sale_report
            expenses.save()

    return HttpResponseRedirect('/accept')


@login_required(login_url='login')
def get_sale_reports(request):
    # просмотр актов продаж
    if 'id' not in request.GET:
        # выводим список актов
        sale_reports = SaleReport.objects.all()

        context = {
            'sale_reports': sale_reports,
        }

        return render(request, 'main/sale_reports.html', context)

    else:
        # выводим статистику по конкретному атку
        policies = Policy.objects.filter(sale_report=request.GET['id'])
        expenses = Expenses.objects.filter(sale_report=request.GET['id'])

        sum_for_final_statistic = {}
        for ex in expenses:
            if ex.name in sum_for_final_statistic:
                sum_for_final_statistic[ex.name] += ex.value
            else:
                sum_for_final_statistic[ex.name] = ex.value

        sum_expenses = round(sum(ex.value for ex in expenses), 2)
        sum_sp = round(sum(policy.sp for policy in policies), 2)
        income = round(sum(policy.sp*(policy.commission/100) for policy in policies), 2)
        profit = round((income - sum_expenses), 2)

        user_statistic = {}
        user_ids = set(policy.user.id for policy in policies)
        for user_id in user_ids:
            user_name = f'{User.objects.get(id=user_id).last_name} {User.objects.get(id=user_id).first_name}'
            temp_dict = {}
            policy_for_user = policies.filter(user_id=user_id)
            temp_dict['user_income'] = round(sum(policy.sp*(policy.commission/100) for policy in policy_for_user), 2)
            temp_dict['osago_count'] = len([policy for policy in policy_for_user if policy.type.name == 'ОСАГО'])
            temp_dict['osago_sum'] = round(
                sum(policy.sp for policy in policy_for_user if policy.type.name == 'ОСАГО'), 2)
            temp_dict['kasko_count'] = len([policy for policy in policy_for_user if policy.type.name == 'КАСКО'])
            temp_dict['kasko_sum'] = round(
                sum(policy.sp for policy in policy_for_user if policy.type.name == 'КАСКО'), 2)
            temp_dict['another_count'] = len(policy_for_user) - temp_dict['osago_count'] - temp_dict['kasko_count']
            temp_dict['another_sum'] = round(
                (sum(policy.sp for policy in policy_for_user) - temp_dict['osago_sum'] - temp_dict['kasko_sum']), 2)
            user_statistic[user_name] = temp_dict

        # ссылка с параметрами для пагинации
        link = '?'
        for key, value in request.GET.items():
            if key == 'page':
                continue
            link = link + f'{key}={value}&'

        paginator = Paginator(policies, 10)
        current_page = request.GET.get('page', 1)
        page = paginator.get_page(current_page)

        context = {
            'sum_for_final_statistic': sum_for_final_statistic,
            'user_statistic': user_statistic,
            'sum_expenses': sum_expenses,
            'sum_sp': sum_sp,
            'profit': profit,
            'income': income,
            'policies': page.object_list,
            'paginator': paginator,
            'page': page,
            'link': link,
        }

        return render(request, 'main/sale_reports_one.html', context)


@login_required(login_url='login')
def base_upload(request):
    pass


@login_required(login_url='login')
def search(request):
    # общий поиск по базе
    if request.user.admin:
        selected = {}
        result = {}
        policy_list = []

        type = Type.objects.all()
        company = Company.objects.all()
        channel = Channel.objects.all()
        users = User.objects.all()

        if request.method == 'POST':
            # вернуть полис в сверки
            if request.method == 'POST' and request.user.admin and 'policy_id_for_return' in request.POST:
                policy_for_return = Policy.objects.get(id=request.POST.get('policy_id_for_return'))
                policy_for_return.accept = False
                policy_for_return.save()

            # удалить полис
            if 'policy_id_for_delete' in request.POST and request.user.admin:
                Policy.objects.get(id=request.POST.get('policy_id_for_delete')).delete()

            # провести полис
            if 'id_policy_for_accept' in request.POST:
                policy_for_accept = Policy.objects.get(id=request.POST.get('id_policy_for_accept'))
                policy_for_accept.accept = True
                policy_for_accept.save()

            # редактировать полис
            if 'id_policy_for_edit' in request.POST:
                policy_for_edit = Policy.objects.get(id=request.POST.get('id_policy_for_edit'))

                full_name = request.POST['full_name_client']
                if f'{policy_for_edit.client.last_name} {policy_for_edit.client.first_name} ' \
                   f'{policy_for_edit.client.middle_name}' != full_name:

                    client_for_edit = policy_for_edit.client

                    while full_name[-1] == ' ':
                        full_name = full_name[:-1]
                    last_name = full_name.split()[0].capitalize()
                    first_name = full_name.split()[1].capitalize()
                    middle_name = ''
                    if len(full_name.split()) > 2:
                        for name in full_name.split()[2:]:
                            middle_name = middle_name + name.capitalize() + " "
                        middle_name = middle_name[:-1]

                    client_for_edit.last_name = last_name
                    client_for_edit.first_name = first_name
                    client_for_edit.middle_name = middle_name
                    client_for_edit.save()

                policy_for_edit.type = Type.objects.get(id=request.POST.get('type'))
                policy_for_edit.status = request.POST.get('Тип_продажи')
                policy_for_edit.number = request.POST.get('number')
                policy_for_edit.series = request.POST.get('series')
                policy_for_edit.company = Company.objects.get(id=request.POST.get('company'))
                policy_for_edit.channel = Channel.objects.get(id=request.POST.get('channel'))
                policy_for_edit.sp = float(request.POST.get('sp').replace(',', '.'))
                policy_for_edit.commission = float(request.POST.get('commission').replace(',', '.'))
                policy_for_edit.date_registration = request.POST.get('date_registration')
                policy_for_edit.date_start = request.POST.get('date_start')
                policy_for_edit.date_end = request.POST.get('date_end')
                policy_for_edit.user = User.objects.get(id=request.POST.get('user'))
                policy_for_edit.save()

        if 'search' in request.GET:
            # поиск по страхователю или по номеру полиса
            selected['search'] = request.GET['search']
            search_list = request.GET['search']
            search_0 = ''
            search_1 = ''
            search_2 = ''
            if search_list != '':
                while search_list[-1] == " ":
                    search_list = search_list[:-1]
                    if len(search_list) == 0:
                        break
                if len(search_list) != 0:
                    while search_list[0] == " ":
                        search_list = search_list[1:]
                        if len(search_list) == 0:
                            break
                if len(search_list) != 0:
                    while "  " in search_list:
                        search_list = search_list.replace("  ", " ")
                    search_list = search_list.split(" ")
                    if len(search_list) > 1:
                        search_0 = search_list[0]
                        search_1 = search_list[1]
                        if len(search_list) > 2:
                            for name in search_list[2:]:
                                search_2 = search_2 + name + " "
                            search_2 = search_2[:-1]
            if len(search_list) != 0:
                result = Policy.objects.filter(
                    Q(client__last_name__iregex=request.GET.get('search')) |
                    Q(client__first_name__iregex=request.GET.get('search')) |
                    Q(client__middle_name__iregex=request.GET.get('search')) |
                    Q(number__iregex=request.GET.get('search')) |
                    Q(series__iregex=request.GET.get('search'))
                    )

                if len(result) == 0 and search_0 != '':
                    if search_2 != '':
                        result = Policy.objects.filter(
                            Q(client__last_name__iregex=search_0) &
                            Q(client__first_name__iregex=search_1) &
                            Q(client__middle_name__iregex=search_2)
                        )
                    elif search_1 != '':
                        result = Policy.objects.filter(
                            Q(client__last_name__iregex=search_0) &
                            Q(client__first_name__iregex=search_1)
                        )

        if len(result) > 0:
            for policy in result:
                commission_rur = '{:.2f}'.format(policy.commission / 100 * policy.sp)
                temp_dict = {
                    'status': policy.status,
                    'accept': policy.accept,
                    'sale_report': policy.sale_report,
                    'type': policy.type,
                    'series': policy.series,
                    'number': policy.number,
                    'company': policy.company,
                    'channel': policy.channel,
                    'sp': policy.sp,
                    'commission': policy.commission,
                    'commission_rur': commission_rur,
                    'client': policy.client,
                    'user': policy.user,
                    'date_registration': policy.date_registration,
                    'date_registration_for_edit': policy.date_registration.strftime("%Y-%m-%d"),
                    'date_start': policy.date_start,
                    'date_start_for_edit': policy.date_start.strftime("%Y-%m-%d"),
                    'date_end': policy.date_end,
                    'date_end_for_edit': policy.date_end.strftime("%Y-%m-%d"),
                    'id': policy.id,
                }
                policy_list.append(temp_dict)

        # ссылка с параметрами для пагинации
        link = '?'
        for key, value in request.GET.items():
            if key == 'page':
                continue
            link = link + f'{key}={value}&'

        paginator = Paginator(policy_list, 20)
        current_page = request.GET.get('page', 1)
        page = paginator.get_page(current_page)

        context = {
            'selected': selected,
            'result': result,
            'policy_list': page.object_list,
            'users': users,
            'companies': company,
            'channels': channel,
            'types': type,
            'page': page,
            'link': link,
            'paginator': paginator,
        }

        return render(request, 'main/search.html', context)
